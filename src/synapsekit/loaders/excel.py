from __future__ import annotations

import os

from .base import Document


class ExcelLoader:
    """Load an Excel (.xlsx) file, one Document per sheet.

    Each sheet is converted to a text table (tab-separated values).
    """

    def __init__(self, path: str) -> None:
        self._path = path

    def load(self) -> list[Document]:
        if not os.path.exists(self._path):
            raise FileNotFoundError(f"Excel file not found: {self._path}")
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl required: pip install synapsekit[excel]") from None

        wb = load_workbook(self._path, read_only=True, data_only=True)
        docs = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                rows.append("\t".join(cells))
            text = "\n".join(rows)
            docs.append(
                Document(
                    text=text,
                    metadata={"source": self._path, "sheet": sheet_name},
                )
            )
        wb.close()
        return docs
